import csv
import io
import json
import re
import shutil
import subprocess
import tempfile
import zipfile
from dataclasses import dataclass, field
from html.parser import HTMLParser
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple
from xml.etree import ElementTree as ET


DAY_NAMES = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
TEXT_EXTENSIONS = {
    ".txt", ".md", ".csv", ".tsv", ".json", ".html", ".htm", ".xml",
    ".rtf", ".log", ".ics", ".yaml", ".yml",
}
IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".webp", ".bmp", ".gif", ".tif", ".tiff"}
ARCHIVE_EXTENSIONS = {".zip"}
EMPTY_SLOT_WORDS = {"", "-", "--", "na", "n/a", "none", "free", "no class"}
BREAK_SLOT_WORDS = {"break", "lunch", "recess", "interval"}


@dataclass
class ExtractedDocument:
    filename: str
    content_type: str
    text: str
    extractor: str
    warnings: List[str] = field(default_factory=list)


def extract_document_text(
    filename: str,
    content: bytes,
    content_type: Optional[str] = None,
    *,
    max_chars: int = 200_000,
    ocr_max_pages: int = 6,
) -> ExtractedDocument:
    """Extract readable text from an uploaded timetable document locally."""
    safe_filename = filename or "uploaded-file"
    ext = Path(safe_filename).suffix.lower()
    warnings: List[str] = []

    if not content:
        return ExtractedDocument(safe_filename, content_type or "", "", "empty", ["File is empty."])

    try:
        if ext == ".pdf":
            text, extra, extractor = _extract_pdf_text(content, ocr_max_pages)
            warnings.extend(extra)
        elif ext == ".docx":
            text = _extract_docx_text(content)
            extractor = "docx"
        elif ext in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
            text, extra = _extract_xlsx_text(content)
            warnings.extend(extra)
            extractor = "xlsx"
        elif ext == ".pptx":
            text = _extract_pptx_text(content)
            extractor = "pptx"
        elif ext in {".odt", ".ods", ".odp"}:
            text = _extract_odf_text(content)
            extractor = "odf"
        elif ext in {".doc", ".xls", ".ppt"}:
            text, extra, extractor = _extract_legacy_office_text(content, ext)
            warnings.extend(extra)
        elif ext in IMAGE_EXTENSIONS:
            text, extra = _extract_image_ocr(content)
            warnings.extend(extra)
            extractor = "ocr-image"
        elif ext in ARCHIVE_EXTENSIONS:
            text, extra = _extract_zip_text(content, max_chars=max_chars, ocr_max_pages=ocr_max_pages)
            warnings.extend(extra)
            extractor = "zip"
        elif ext in TEXT_EXTENSIONS or _looks_like_text(content):
            text = _extract_text_like(content, ext)
            extractor = "text"
        else:
            text = _decode_bytes(content)
            extractor = "best-effort"
            if not _looks_like_text(content):
                warnings.append(
                    "Unknown binary format. Used best-effort text decoding; install a format-specific extractor if important text is missing."
                )
    except Exception as exc:
        text = _decode_bytes(content)
        extractor = "fallback"
        warnings.append(f"Primary extractor failed for {safe_filename}: {exc}. Used best-effort decoding.")

    text = _clean_text(text)
    if len(text) > max_chars:
        text = text[:max_chars]
        warnings.append(f"Text was truncated to {max_chars:,} characters for parsing.")

    if not text.strip():
        warnings.append(
            "No readable text was found. If this is a scanned PDF or image, install Tesseract OCR and Poppler locally."
        )

    return ExtractedDocument(safe_filename, content_type or "", text, extractor, warnings)


def build_constraint_text_from_documents(
    documents: Iterable[ExtractedDocument],
    context: Dict[str, Any],
) -> Tuple[str, List[Dict[str, Any]], List[str]]:
    """Convert extracted timetable text into natural-language fixed-slot constraints."""
    constraints: List[Dict[str, Any]] = []
    warnings: List[str] = []

    for document in documents:
        if not document.text.strip():
            continue
        json_constraints = _constraints_from_json_text(document.text)
        if json_constraints:
            constraints.extend(json_constraints)
            continue
        constraints.extend(_constraints_from_table_text(document.text, context))

    constraints = _dedupe_constraints(constraints)
    if constraints:
        return _constraints_to_text(constraints), constraints, warnings

    fallback_text = "\n\n".join(doc.text for doc in documents if doc.text.strip())
    if fallback_text:
        warnings.append("No timetable table was detected, so the extracted text will be parsed directly.")
    return fallback_text, [], warnings


def _extract_text_like(content: bytes, ext: str) -> str:
    text = _decode_bytes(content)
    if ext in {".csv", ".tsv"}:
        return _normalize_delimited_text(text, "\t" if ext == ".tsv" else None)
    if ext == ".json":
        try:
            return json.dumps(json.loads(text), indent=2, ensure_ascii=False)
        except Exception:
            return text
    if ext in {".html", ".htm"}:
        parser = _TextHTMLParser()
        parser.feed(text)
        return parser.text()
    if ext == ".xml":
        return _xml_text(text)
    if ext == ".rtf":
        return _rtf_to_text(text)
    return text


def _extract_pdf_text(content: bytes, ocr_max_pages: int) -> Tuple[str, List[str], str]:
    warnings: List[str] = []
    extractor = "pdf"
    text = ""

    for module_name in ("pypdf", "PyPDF2"):
        try:
            module = __import__(module_name)
            reader = module.PdfReader(io.BytesIO(content))
            pages = []
            for page in reader.pages:
                pages.append(page.extract_text() or "")
            text = "\n".join(pages)
            extractor = module_name
            break
        except Exception as exc:
            warnings.append(f"{module_name} could not read PDF text: {exc}")

    cli_text = _pdftotext_cli(content)
    if cli_text.strip() and (len(text.strip()) < 80 or _looks_more_layout_preserving(cli_text, text)):
        text = cli_text
        extractor = "pdftotext"

    if len(text.strip()) < 80:
        ocr_text, ocr_warnings = _extract_pdf_ocr(content, ocr_max_pages)
        warnings.extend(ocr_warnings)
        if ocr_text.strip():
            text = ocr_text
            extractor = "ocr-pdf"

    return text, warnings, extractor


def _extract_docx_text(content: bytes) -> str:
    with zipfile.ZipFile(io.BytesIO(content)) as archive:
        paths = [
            name for name in archive.namelist()
            if name.startswith("word/") and name.endswith(".xml")
            and any(part in name for part in ("document", "header", "footer"))
        ]
        return "\n".join(_openxml_text(archive.read(path)) for path in paths)


def _extract_xlsx_text(content: bytes) -> Tuple[str, List[str]]:
    warnings: List[str] = []
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        lines: List[str] = []
        for sheet in workbook.worksheets:
            lines.append(f"Sheet: {sheet.title}")
            for row in sheet.iter_rows(values_only=True):
                values = ["" if value is None else str(value).strip() for value in row]
                if any(values):
                    lines.append("\t".join(values))
        return "\n".join(lines), warnings
    except Exception as exc:
        warnings.append(f"openpyxl could not read workbook; using XML fallback: {exc}")

    with zipfile.ZipFile(io.BytesIO(content)) as archive:
        shared_strings = _xlsx_shared_strings(archive)
        sheet_paths = sorted(
            path for path in archive.namelist()
            if path.startswith("xl/worksheets/sheet") and path.endswith(".xml")
        )
        lines: List[str] = []
        for path in sheet_paths:
            lines.append(f"Sheet: {Path(path).stem}")
            lines.extend(_xlsx_sheet_rows(archive.read(path), shared_strings))
        return "\n".join(lines), warnings


def _extract_pptx_text(content: bytes) -> str:
    with zipfile.ZipFile(io.BytesIO(content)) as archive:
        slide_paths = sorted(
            name for name in archive.namelist()
            if name.startswith("ppt/slides/slide") and name.endswith(".xml")
        )
        return "\n".join(_openxml_text(archive.read(path)) for path in slide_paths)


def _extract_odf_text(content: bytes) -> str:
    with zipfile.ZipFile(io.BytesIO(content)) as archive:
        if "content.xml" not in archive.namelist():
            return ""
        return _xml_text(archive.read("content.xml").decode("utf-8", errors="ignore"))


def _extract_legacy_office_text(content: bytes, ext: str) -> Tuple[str, List[str], str]:
    warnings: List[str] = []
    if ext == ".doc" and shutil.which("antiword"):
        text = _run_cli_with_temp(content, ext, ["antiword"])
        if text.strip():
            return text, warnings, "antiword"

    if ext == ".xls" and shutil.which("xls2csv"):
        text = _run_cli_with_temp(content, ext, ["xls2csv"])
        if text.strip():
            return text, warnings, "xls2csv"

    converted = _libreoffice_convert_text(content, ext)
    if converted.strip():
        return converted, warnings, "libreoffice"

    warnings.append(
        f"Legacy {ext} files need LibreOffice, antiword, or xls2csv for accurate extraction."
    )
    return _decode_bytes(content), warnings, "legacy-fallback"


def _extract_zip_text(content: bytes, *, max_chars: int, ocr_max_pages: int) -> Tuple[str, List[str]]:
    warnings: List[str] = []
    parts: List[str] = []
    total_chars = 0
    with zipfile.ZipFile(io.BytesIO(content)) as archive:
        members = [
            name for name in archive.namelist()
            if not name.endswith("/") and not Path(name).name.startswith(".")
        ][:30]
        for member in members:
            if total_chars >= max_chars:
                break
            try:
                data = archive.read(member)
            except Exception as exc:
                warnings.append(f"Skipped {member}: {exc}")
                continue
            extracted = extract_document_text(
                member,
                data,
                max_chars=max_chars - total_chars,
                ocr_max_pages=ocr_max_pages,
            )
            warnings.extend([f"{member}: {warning}" for warning in extracted.warnings])
            if extracted.text.strip():
                part = f"File: {member}\n{extracted.text}"
                parts.append(part)
                total_chars += len(part)
    return "\n\n".join(parts), warnings


def _extract_image_ocr(content: bytes) -> Tuple[str, List[str]]:
    try:
        from PIL import Image
        import pytesseract

        image = Image.open(io.BytesIO(content))
        return pytesseract.image_to_string(image), []
    except Exception as exc:
        return "", [f"Local image OCR is unavailable or failed: {exc}"]


def _extract_pdf_ocr(content: bytes, max_pages: int) -> Tuple[str, List[str]]:
    try:
        from pdf2image import convert_from_bytes
        import pytesseract

        pages = convert_from_bytes(content, first_page=1, last_page=max(1, max_pages))
        text = "\n".join(pytesseract.image_to_string(page) for page in pages)
        return text, []
    except Exception as exc:
        return "", [f"Local PDF OCR is unavailable or failed: {exc}"]


def _pdftotext_cli(content: bytes) -> str:
    if not shutil.which("pdftotext"):
        return ""
    return _run_cli_with_temp(content, ".pdf", ["pdftotext", "-layout", "{input}", "-"])


def _looks_more_layout_preserving(candidate: str, current: str) -> bool:
    if len(candidate.strip()) < max(80, int(len(current.strip()) * 0.6)):
        return False
    candidate_score = sum(1 for line in candidate.splitlines() if "\t" in line or re.search(r"\S\s{2,}\S", line))
    current_score = sum(1 for line in current.splitlines() if "\t" in line or re.search(r"\S\s{2,}\S", line))
    return candidate_score > current_score


def _libreoffice_convert_text(content: bytes, ext: str) -> str:
    binary = shutil.which("libreoffice") or shutil.which("soffice")
    if not binary:
        return ""
    with tempfile.TemporaryDirectory() as tmpdir:
        input_path = Path(tmpdir) / f"input{ext}"
        input_path.write_bytes(content)
        cmd = [
            binary, "--headless", "--convert-to", "txt:Text",
            "--outdir", tmpdir, str(input_path),
        ]
        try:
            subprocess.run(cmd, capture_output=True, timeout=45, check=False)
        except Exception:
            return ""
        for candidate in Path(tmpdir).glob("*.txt"):
            return candidate.read_text(encoding="utf-8", errors="ignore")
    return ""


def _run_cli_with_temp(content: bytes, ext: str, command: List[str]) -> str:
    with tempfile.NamedTemporaryFile(suffix=ext) as handle:
        handle.write(content)
        handle.flush()
        cmd = [part.format(input=handle.name) for part in command]
        if "{input}" not in " ".join(command):
            cmd.append(handle.name)
        try:
            result = subprocess.run(cmd, capture_output=True, timeout=30, check=False)
        except Exception:
            return ""
        return result.stdout.decode("utf-8", errors="ignore")


def _constraints_from_json_text(text: str) -> List[Dict[str, Any]]:
    try:
        data = json.loads(text)
    except Exception:
        return []

    if isinstance(data, dict) and "schedule_data" in data:
        data = data["schedule_data"]
    if isinstance(data, dict) and "timetable" in data:
        data = {"uploaded_class": data}
    if not isinstance(data, dict):
        return []

    constraints: List[Dict[str, Any]] = []
    for class_schedule in data.values():
        if not isinstance(class_schedule, dict):
            continue
        class_name = class_schedule.get("class_name") or class_schedule.get("name")
        timetable = class_schedule.get("timetable")
        if not class_name or not isinstance(timetable, dict):
            continue
        for day, slots in timetable.items():
            if not isinstance(slots, list):
                continue
            for slot in slots:
                if not isinstance(slot, dict) or slot.get("slot_type") == "break":
                    continue
                subject = slot.get("subject") or slot.get("subject_code")
                period = slot.get("period")
                if subject and period:
                    constraints.append(_specific_slot(str(subject), str(class_name), str(day), period))
    return constraints


def _constraints_from_table_text(text: str, context: Dict[str, Any]) -> List[Dict[str, Any]]:
    subject_names = list(context.get("subject_names") or [])
    class_names = list(context.get("class_names") or [])
    lines = [line.strip() for line in text.splitlines() if line.strip()]

    constraints: List[Dict[str, Any]] = []
    active_class: Optional[str] = None
    header_periods: Dict[int, int] = {}

    for line in lines:
        cells = _split_table_cells(line)
        line_day = _find_day(line)
        line_class = _match_name(line, class_names)

        if line_class and _looks_like_class_header(line, line_class, line_day):
            active_class = line_class

        if cells:
            parsed_header = _header_period_map(cells)
            if parsed_header:
                header_periods = parsed_header
                continue

            row_day = _find_day(cells[0]) or line_day
            if row_day and len(cells) > 1:
                row_class = line_class or active_class
                constraints.extend(
                    _constraints_from_table_row(cells, row_day, row_class, header_periods, subject_names)
                )
                continue

        if line_day:
            line_subjects = _match_names(line, subject_names)
            row_class = line_class or active_class
            periods = _period_numbers(line)
            if line_subjects and periods:
                for subject in line_subjects[:2]:
                    for period in periods:
                        constraints.append(_specific_slot(subject, row_class, line_day, period))

    return constraints


def _constraints_from_table_row(
    cells: List[str],
    day: str,
    class_name: Optional[str],
    header_periods: Dict[int, int],
    subject_names: List[str],
) -> List[Dict[str, Any]]:
    constraints: List[Dict[str, Any]] = []
    fallback_period = 0
    for index, cell in enumerate(cells[1:], start=1):
        period = header_periods.get(index)
        if period is None:
            if _is_break_slot(cell):
                continue
            fallback_period += 1
            period = fallback_period
        if _is_empty_slot(cell):
            continue
        subject = _match_name(cell, subject_names)
        if not subject:
            continue
        constraints.append(_specific_slot(subject, class_name, day, period))
    return constraints


def _specific_slot(subject: str, class_name: Optional[str], day: str, period: Any) -> Dict[str, Any]:
    constraint: Dict[str, Any] = {
        "type": "specific_time_slot",
        "target": str(subject),
        "target_type": "subject",
        "day": _canonical_day(day) or str(day),
        "period": int(period),
        "hard": True,
    }
    if class_name:
        constraint["class_name"] = str(class_name)
    return constraint


def _constraints_to_text(constraints: List[Dict[str, Any]]) -> str:
    lines = []
    for c in constraints:
        prefix = f"For Class {c['class_name']}, " if c.get("class_name") else ""
        lines.append(f"{prefix}{c['target']} must be on {c['day']} period {c['period']}.")
    return "\n".join(lines)


def _dedupe_constraints(constraints: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    deduped: List[Dict[str, Any]] = []
    seen = set()
    for constraint in constraints:
        key = (
            constraint.get("type"),
            _normalize_name(constraint.get("target")),
            _normalize_name(constraint.get("class_name")),
            _canonical_day(constraint.get("day")),
            int(constraint.get("period", 0)),
        )
        if key in seen:
            continue
        seen.add(key)
        deduped.append(constraint)
    return deduped


def _decode_bytes(content: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-16", "cp1252", "latin-1"):
        try:
            decoded = content.decode(encoding)
            if _printable_ratio(decoded) >= 0.72:
                return decoded
        except Exception:
            continue
    return content.decode("utf-8", errors="ignore")


def _looks_like_text(content: bytes) -> bool:
    sample = content[:4096]
    if not sample:
        return True
    if b"\x00" in sample:
        return False
    decoded = sample.decode("utf-8", errors="ignore")
    return _printable_ratio(decoded) >= 0.72


def _printable_ratio(text: str) -> float:
    if not text:
        return 1.0
    printable = sum(1 for ch in text if ch.isprintable() or ch in "\n\r\t")
    return printable / max(len(text), 1)


def _clean_text(text: str) -> str:
    text = (text or "").replace("\x00", " ")
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    lines = []
    for line in text.split("\n"):
        line = re.sub(r"[ \f\v]+", " ", line).strip(" ")
        if line.strip():
            lines.append(line)
    return "\n".join(lines)


def _normalize_delimited_text(text: str, delimiter: Optional[str]) -> str:
    dialect_delimiter = delimiter
    if dialect_delimiter is None:
        try:
            dialect_delimiter = csv.Sniffer().sniff(text[:4096]).delimiter
        except Exception:
            dialect_delimiter = ","
    reader = csv.reader(io.StringIO(text), delimiter=dialect_delimiter)
    return "\n".join("\t".join(cell.strip() for cell in row) for row in reader)


class _TextHTMLParser(HTMLParser):
    block_tags = {"p", "div", "tr", "table", "section", "article", "li", "h1", "h2", "h3", "h4"}

    def __init__(self):
        super().__init__()
        self.parts: List[str] = []

    def handle_starttag(self, tag: str, attrs: List[Tuple[str, Optional[str]]]):
        if tag in self.block_tags:
            self.parts.append("\n")
        elif tag in {"td", "th"}:
            self.parts.append("\t")
        elif tag == "br":
            self.parts.append("\n")

    def handle_data(self, data: str):
        if data.strip():
            self.parts.append(data)

    def text(self) -> str:
        return "".join(self.parts)


def _xml_text(text: str) -> str:
    try:
        root = ET.fromstring(text)
        return " ".join(part.strip() for part in root.itertext() if part and part.strip())
    except Exception:
        return re.sub(r"<[^>]+>", " ", text)


def _rtf_to_text(text: str) -> str:
    text = re.sub(
        r"\\'([0-9a-fA-F]{2})",
        lambda match: bytes.fromhex(match.group(1)).decode("latin-1", errors="ignore"),
        text,
    )
    text = re.sub(r"\\[a-zA-Z]+-?\d* ?", " ", text)
    text = text.replace("\\~", " ").replace("\\_", "-")
    return re.sub(r"[{}]", " ", text)


def _openxml_text(xml_bytes: bytes) -> str:
    try:
        root = ET.fromstring(xml_bytes)
    except Exception:
        return ""

    parts: List[str] = []

    def walk(node: ET.Element):
        tag = node.tag.rsplit("}", 1)[-1]
        if tag in {"t", "instrText"} and node.text:
            parts.append(node.text)
        elif tag == "tab":
            parts.append("\t")
        for child in node:
            walk(child)
        if tag in {"p", "tr"}:
            parts.append("\n")
        elif tag == "tc":
            parts.append("\t")

    walk(root)
    return "".join(parts)


def _xlsx_shared_strings(archive: zipfile.ZipFile) -> List[str]:
    if "xl/sharedStrings.xml" not in archive.namelist():
        return []
    root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
    strings = []
    for item in root:
        strings.append("".join(item.itertext()))
    return strings


def _xlsx_sheet_rows(xml_bytes: bytes, shared_strings: List[str]) -> List[str]:
    root = ET.fromstring(xml_bytes)
    rows: List[str] = []
    for row in root.iter():
        if row.tag.rsplit("}", 1)[-1] != "row":
            continue
        values: List[str] = []
        current_col = 1
        for cell in row:
            if cell.tag.rsplit("}", 1)[-1] != "c":
                continue
            col = _column_number(cell.attrib.get("r", ""))
            while col and current_col < col:
                values.append("")
                current_col += 1
            values.append(_xlsx_cell_value(cell, shared_strings))
            current_col += 1
        if any(value.strip() for value in values):
            rows.append("\t".join(values))
    return rows


def _xlsx_cell_value(cell: ET.Element, shared_strings: List[str]) -> str:
    cell_type = cell.attrib.get("t")
    if cell_type == "inlineStr":
        return "".join(cell.itertext()).strip()
    value = ""
    for child in cell:
        if child.tag.rsplit("}", 1)[-1] == "v" and child.text is not None:
            value = child.text
            break
    if cell_type == "s":
        try:
            return shared_strings[int(value)].strip()
        except Exception:
            return ""
    return str(value).strip()


def _column_number(cell_ref: str) -> Optional[int]:
    match = re.match(r"([A-Z]+)", cell_ref or "")
    if not match:
        return None
    total = 0
    for char in match.group(1):
        total = total * 26 + (ord(char) - ord("A") + 1)
    return total


def _split_table_cells(line: str) -> List[str]:
    if "\t" in line:
        return [cell.strip() for cell in line.split("\t")]
    if "|" in line and line.count("|") >= 2:
        return [cell.strip() for cell in line.strip("|").split("|")]
    if re.search(r" {2,}", line):
        return [cell.strip() for cell in re.split(r" {2,}", line) if cell.strip()]
    return []


def _header_period_map(cells: List[str]) -> Dict[int, int]:
    mapping: Dict[int, int] = {}
    for index, cell in enumerate(cells):
        lower = cell.lower()
        if "day" in lower and index == 0:
            continue
        match = re.search(r"\b(?:period|p)\s*#?\s*(\d{1,2})\b", lower)
        if match:
            mapping[index] = int(match.group(1))
    return mapping


def _period_numbers(text: str) -> List[int]:
    periods = [
        int(match.group(1))
        for match in re.finditer(r"\b(?:period|p)\s*#?\s*(\d{1,2})\b", text, flags=re.IGNORECASE)
        if 1 <= int(match.group(1)) <= 20
    ]
    if periods:
        return sorted(set(periods))
    if _find_day(text):
        return sorted({
            int(raw)
            for raw in re.findall(r"\b(\d{1,2})\b", text)
            if 1 <= int(raw) <= 20
        })
    return []


def _find_day(text: Any) -> Optional[str]:
    normalized = _normalize_name(text)
    for day in DAY_NAMES:
        if re.search(r"\b" + re.escape(day.lower()) + r"\b", normalized):
            return day
    aliases = {
        "mon": "Monday", "tue": "Tuesday", "wed": "Wednesday", "thu": "Thursday",
        "fri": "Friday", "sat": "Saturday", "sun": "Sunday",
    }
    for alias, day in aliases.items():
        if re.search(r"\b" + alias + r"\b", normalized):
            return day
    return None


def _canonical_day(value: Any) -> Optional[str]:
    return _find_day(value)


def _match_names(text: Any, names: List[str]) -> List[str]:
    normalized = _normalize_name(text)
    matches: List[Tuple[int, str]] = []
    occupied: set = set()
    for name in sorted((n for n in names if n), key=lambda n: len(_normalize_name(n)), reverse=True):
        needle = _normalize_name(name)
        if not needle:
            continue
        match = re.search(r"\b" + re.escape(needle) + r"\b", normalized)
        if not match:
            continue
        span = set(range(match.start(), match.end()))
        if occupied & span:
            continue
        occupied |= span
        matches.append((match.start(), name))
    return [name for _, name in sorted(matches)]


def _match_name(text: Any, names: List[str]) -> Optional[str]:
    matches = _match_names(text, names)
    return matches[0] if matches else None


def _normalize_name(value: Any) -> str:
    text = str(value or "").strip()
    text = re.sub(r"([a-z])([A-Z])", r"\1 \2", text)
    text = text.lower()
    return re.sub(r"[^a-z0-9]+", " ", text).strip()


def _looks_like_class_header(line: str, class_name: str, day: Optional[str]) -> bool:
    if day:
        return False
    normalized_line = _normalize_name(line)
    normalized_class = _normalize_name(class_name)
    if normalized_line == normalized_class:
        return True
    return any(word in normalized_line for word in ("class", "section", "timetable", "schedule"))


def _is_empty_slot(cell: str) -> bool:
    normalized = _normalize_name(cell)
    if normalized in EMPTY_SLOT_WORDS:
        return True
    return _is_break_slot(cell)


def _is_break_slot(cell: str) -> bool:
    normalized = _normalize_name(cell)
    return normalized in BREAK_SLOT_WORDS or any(word in normalized.split() for word in BREAK_SLOT_WORDS)
